"""Rebuild the total-games calibration (MEAN_BY_CLOSENESS + GAMES_CURVE) from the
tennisdata xlsx files. Paste the printed tables into props/games_model.py.

    PYTHONPATH=src .venv/bin/python scripts/build_games_calibration.py

total games = sum of both players' games across all completed sets; predictor =
match competitiveness (de-vigged PSW/PSL) + best_of. See props/games_model.py.
"""
from __future__ import annotations

import glob
import statistics
from collections import defaultdict

from openpyxl import load_workbook

from tennis_wc.config import get_settings  # noqa: F401  (ensures src on path when run via module)

_EXT = "src/data/external/tennisdata_*.xlsx"


def _load() -> list[tuple[float, float, int]]:
    recs = []
    for path in sorted(glob.glob(_EXT)):
        wb = load_workbook(path, read_only=True); ws = wb.active
        hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(hdr)}
        if any(n not in idx for n in ("Best of", "Comment", "PSW", "PSL", "W1", "L1")):
            continue
        for r in ws.iter_rows(min_row=2, values_only=True):
            try:
                if str(r[idx["Comment"]]).strip() != "Completed":
                    continue
                bo = int(r[idx["Best of"]])
                psw, psl = float(r[idx["PSW"]]), float(r[idx["PSL"]])
            except (TypeError, ValueError):
                continue
            tg, ok = 0, False
            for s in range(1, 6):
                wi, li = idx.get(f"W{s}"), idx.get(f"L{s}")
                if wi is None or li is None:
                    continue
                w, l = r[wi], r[li]
                if w is None or l is None:
                    continue
                try:
                    tg += int(w) + int(l); ok = True
                except (TypeError, ValueError):
                    pass
            if not ok or tg < 6 or psw <= 1 or psl <= 1:
                continue
            p = (1 / psw) / (1 / psw + 1 / psl)
            recs.append((tg, 1.0 - abs(2 * p - 1), bo))
    return recs


def main() -> None:
    recs = _load()
    print(f"# {len(recs)} completed matches")
    mean_by = defaultdict(list)
    for tg, cl, bo in recs:
        mean_by[(bo, round(cl * 5) / 5)].append(tg)
    mean_tbl = {k: round(statistics.mean(v), 2) for k, v in mean_by.items() if len(v) >= 50}
    print("MEAN_BY_CLOSENESS =", dict(sorted(mean_tbl.items())))

    def pm(cl, bo):
        return mean_tbl.get((bo, round(cl * 5) / 5)) or mean_tbl.get((3, 0.6)) or 21.0

    bins = defaultdict(lambda: [0, 0])
    for tg, cl, bo in recs:
        m = pm(cl, bo)
        for line in [x + 0.5 for x in range(12, 40)]:
            b = round((line / m) * 20) / 20
            bins[b][0] += 1 if tg > line else 0
            bins[b][1] += 1
    curve = [(b, round(w / n, 4)) for b, (w, n) in sorted(bins.items()) if n >= 300 and 0.6 <= b <= 1.45]
    print("GAMES_CURVE =", curve)


if __name__ == "__main__":
    main()
